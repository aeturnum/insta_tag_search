import uuid
from time import time
from os.path import join
from pathlib import Path

from openpyxl import Workbook, worksheet

# from constants import KEEP_XLSX_FOR

class WorkbookManager(object):
    _user_labels = [
        'Username', 'Display Name', 'Profile Picture URL'
    ]
    _post_labels = [
        'Global ID', 'Username', 'Timestamp', 'Location', 'Comments', 'Image URL', 'Likes',
        'Caption'
    ]
    _sheet_info = ['Tag', 'Account Used']

    _sheet_path = './'

    def __init__(self, tag, username):
        wb = Workbook()
        self._name = f'{str(uuid.uuid4())}.xlsx'
        self._wb = wb
        # title page that has the meta info
        main = WorksheetManager(wb.active)
        main.title = "Comment Source"
        main.append(self._sheet_info)
        main.append([tag, username])

        self.post_worksheet = WorksheetManager(wb.create_sheet("Posts"))
        self.post_worksheet.append(self._post_labels)

        self.user_worksheet = WorksheetManager(wb.create_sheet("Users"))
        self.user_worksheet.append(self._user_labels)

    def _path(self, file:str):
        return join(self._sheet_path, file)

    @property
    def path(self):
        return self._path(self._name)

    # def _cleanup(self):
    #     # check to see if we
    #     now = time()
    #     directory = Path(self._sheet_path)
    #     for xlsx in directory.iterdir():
    #         if now - xlsx.stat().st_ctime > KEEP_XLSX_FOR:
    #             xlsx.unlink()


    def save(self):
        # self._cleanup()
        self._wb.save(self.path)

class WorksheetManager(object):
    """
    Helper class to track the size of the various columns and create a readable worksheet
    """
    # if we need more we can add them
    _col_letters = ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'M', 'O', 'P')

    def __init__(self, workbook: worksheet.worksheet.Worksheet):
        self._wb = workbook
        self._widths = {}
        self._members = set()

    @property
    def title(self):
        return self._wb.title

    @title.setter
    def title(self, value):
        self._wb.title = value

    def append(self, row, key=None):
        if key is not None:
            if key in self._members:
                # skip
                return
            self._members.add(key)
        for idx, value in enumerate(row):
            # todo: take newlines into account
            length = len(str(value))
            letter = self._col_letters[idx]
            if length > self._widths.get(letter, 0):
                self._widths[letter] = length
                self._wb.column_dimensions[letter].width = length

        self._wb.append(row)
